import datetime
import openpyxl

# открываем файл Excel
workbook = openpyxl.load_workbook('example.xlsx')

# выбираем лист с текущим месяцем
month = datetime.datetime.now().month
worksheet = workbook[str(month)]

# ищем строку с текущей датой
today = datetime.datetime.now().date()
temp_first, temp_second, temp_main = None, None, None
for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
    if row[0].value == today:
        temp_first = row[0].value
        temp_second = row[1].value
        temp_main = row[2].value
        break

# ищем значения temp_second и temp_main на 13 листе
if temp_second is not None and temp_main is not None:
    sheet13 = workbook['13']
    duty_second, duty_main = None, None
    for row in sheet13.iter_rows(min_row=1, max_row=sheet13.max_row):
        if row[0].value == temp_second:
            duty_second = row[1].value
        if row[0].value == temp_main:
            duty_main = row[1].value
        if duty_second is not None and duty_main is not None:
            break

# создаем список и заполняем его в соответствии с текущим временем
duty_all = []
current_time = datetime.datetime.now().time()
if datetime.time(8, 0) <= current_time <= datetime.time(17, 0):
    duty_all.append(duty_main)
if datetime.time(11, 0) <= current_time <= datetime.time(20, 0):
    duty_all.append(duty_first)
if current_time >= datetime.time(20, 0) or current_time < datetime.time(8, 0):
    duty_all.append(duty_second)

# выводим результаты
print(f"duty_all: {duty_all}")
